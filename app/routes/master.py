from flask import Blueprint, render_template, request, redirect, flash, url_for, Response
from app.models.models import Examen, ExamenAlumno, Alumno, Estacion, Categoria, Criterio, Evaluacion, EvaluacionDetalle
from extensions import db
import fcntl
import os
import socket
import struct
import csv
import io
import re
import unicodedata
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.exc import IntegrityError

master_bp = Blueprint('master', __name__)
XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def get_active_exam(create_if_missing=True):
    examen = Examen.query.filter_by(estado='activo').order_by(Examen.id.desc()).first()
    if examen or not create_if_missing:
        return examen

    examen = Examen(nombre='Examen activo', descripcion='Creado automáticamente por TeleECOE')
    db.session.add(examen)
    db.session.commit()
    return examen

def enroll_student(examen, alumno, observaciones=None):
    if not examen or not alumno:
        return None
    existente = ExamenAlumno.query.filter_by(examen_id=examen.id, alumno_id=alumno.id).first()
    if existente:
        return existente
    item = ExamenAlumno(examen_id=examen.id, alumno_id=alumno.id, observaciones=observaciones)
    db.session.add(item)
    return item

def get_interface_ip(interface_name):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            request_data = struct.pack('256s', interface_name[:15].encode('utf-8'))
            return socket.inet_ntoa(fcntl.ioctl(s.fileno(), 0x8915, request_data)[20:24])
    except OSError:
        return None

def get_local_ip():
    configured_host = os.environ.get('TABLET_HOST')
    if configured_host:
        return configured_host

    hotspot_ip = get_interface_ip('ap0')
    if hotspot_ip:
        return hotspot_ip

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

def get_tablet_url():
    configured_url = os.environ.get('TABLET_URL')
    if configured_url:
        return configured_url

    return f'http://{get_local_ip()}:5000/tablet/'

def slugify_identifier(value):
    value = (value or '').strip().lower()
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^a-z0-9]+', '_', value)
    value = re.sub(r'_+', '_', value).strip('_')
    return value[:40]

def unique_criterio_id(base_id):
    base_id = slugify_identifier(base_id) or 'criterio'
    candidate = base_id
    suffix = 2
    while Criterio.query.get(candidate):
        suffix_text = f'_{suffix}'
        candidate = f'{base_id[:50 - len(suffix_text)]}{suffix_text}'
        suffix += 1
    return candidate

def style_export_sheet(ws):
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    for column_cells in ws.columns:
        width = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 55)
    ws.freeze_panes = 'A2'

@master_bp.route('/')
def dashboard():
    examen_activo = get_active_exam()
    alumnos = (
        Alumno.query
        .join(ExamenAlumno)
        .filter(ExamenAlumno.examen_id == examen_activo.id)
        .order_by(Alumno.grupo, Alumno.nombre)
        .all()
    )
    estaciones = Estacion.query.filter_by(activa=True).order_by(Estacion.orden).all()
    
    resultados = []
    for alu in alumnos:
        notas_estaciones = {}
        evaluaciones = {
            e.estacion_id: e
            for e in alu.evaluaciones
            if e.examen_id == examen_activo.id
        }
        for est in estaciones:
            eval_record = evaluaciones.get(est.id)
            if eval_record:
                notas_estaciones[est.id] = {'estado': True, 'puntaje': eval_record.puntaje_total}
            else:
                notas_estaciones[est.id] = {'estado': False, 'puntaje': 0.0}
        
        resultados.append({
            'id': alu.id,
            'nombre': alu.nombre,
            'grupo': alu.grupo,
            'notas_estaciones': notas_estaciones
        })
        
    tablet_url = get_tablet_url()
    return render_template(
        'master_dashboard.html',
        ip=get_local_ip(),
        tablet_url=tablet_url,
        estaciones=estaciones,
        resultados=resultados,
        examen_activo=examen_activo
    )

@master_bp.route('/tablet-qr.svg')
def tablet_qr():
    import qrcode
    from qrcode.image.svg import SvgPathImage

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2,
        image_factory=SvgPathImage,
    )
    qr.add_data(get_tablet_url())
    qr.make(fit=True)

    output = io.BytesIO()
    qr.make_image().save(output)
    return Response(
        output.getvalue(),
        mimetype='image/svg+xml',
        headers={'Cache-Control': 'no-store'}
    )

@master_bp.route('/examenes', methods=['GET', 'POST'])
def examenes_admin():
    if request.method == 'POST':
        accion = request.form.get('accion')

        if accion == 'crear':
            nombre = (request.form.get('nombre') or '').strip()
            descripcion = (request.form.get('descripcion') or '').strip() or None
            copiar = request.form.get('copiar_alumnos') == '1'
            if not nombre:
                flash('El nombre del examen es obligatorio.', 'danger')
                return redirect(url_for('master.examenes_admin'))

            actual = get_active_exam(create_if_missing=False)
            if actual:
                actual.estado = 'cerrado'
                actual.fecha_cierre = actual.fecha_cierre or datetime.utcnow()

            nuevo = Examen(nombre=nombre, descripcion=descripcion, estado='activo')
            db.session.add(nuevo)
            db.session.flush()

            if copiar and actual:
                for inscripcion in ExamenAlumno.query.filter_by(examen_id=actual.id).all():
                    enroll_student(nuevo, inscripcion.alumno)

            db.session.commit()
            flash('Nuevo examen creado y activado.', 'success')
            return redirect(url_for('master.dashboard'))

        if accion == 'cerrar':
            examen = get_active_exam(create_if_missing=False)
            if examen:
                examen.estado = 'cerrado'
                examen.fecha_cierre = datetime.utcnow()
                db.session.commit()
                flash('Examen cerrado. Puedes crear uno nuevo cuando lo necesites.', 'success')
            return redirect(url_for('master.examenes_admin'))

        if accion == 'activar':
            examen_id = request.form.get('examen_id', type=int)
            examen = Examen.query.get_or_404(examen_id)
            Examen.query.filter_by(estado='activo').update({'estado': 'cerrado', 'fecha_cierre': datetime.utcnow()})
            examen.estado = 'activo'
            examen.fecha_cierre = None
            db.session.commit()
            flash(f'Examen activo cambiado a: {examen.nombre}', 'success')
            return redirect(url_for('master.dashboard'))

        if accion == 'eliminar':
            examen_id = request.form.get('examen_id', type=int)
            confirmacion = (request.form.get('confirmacion') or '').strip().upper()
            examen = Examen.query.get_or_404(examen_id)

            if examen.estado == 'activo':
                flash('No se puede eliminar el examen activo. Activa otro examen o ciérralo antes de eliminarlo.', 'danger')
                return redirect(url_for('master.examenes_admin'))

            if confirmacion != 'ELIMINAR':
                flash('El examen no fue eliminado porque la confirmación no coincidió.', 'warning')
                return redirect(url_for('master.examenes_admin'))

            nombre = examen.nombre
            evaluaciones = Evaluacion.query.filter_by(examen_id=examen.id).all()
            evaluacion_ids = [e.id for e in evaluaciones]
            if evaluacion_ids:
                EvaluacionDetalle.query.filter(EvaluacionDetalle.evaluacion_id.in_(evaluacion_ids)).delete(synchronize_session=False)
                Evaluacion.query.filter(Evaluacion.id.in_(evaluacion_ids)).delete(synchronize_session=False)
            ExamenAlumno.query.filter_by(examen_id=examen.id).delete(synchronize_session=False)
            db.session.delete(examen)
            db.session.commit()
            flash(f'Examen eliminado: {nombre}. También se eliminaron sus evaluaciones asociadas.', 'success')
            return redirect(url_for('master.examenes_admin'))

    examenes = Examen.query.order_by(Examen.id.desc()).all()
    examen_activo = get_active_exam(create_if_missing=False)
    return render_template('master_examenes.html', examenes=examenes, examen_activo=examen_activo)

@master_bp.route('/alumnos', methods=['GET', 'POST'])
def alumnos_admin():
    examen_activo = get_active_exam()
    if request.method == 'POST':
        id_val = (request.form.get('id') or '').strip()
        nombre = (request.form.get('nombre') or '').strip()
        cmp = (request.form.get('cmp') or '').strip() or None
        grupo = (request.form.get('grupo') or '').strip()

        if not nombre:
            flash('El nombre del alumno es obligatorio.', 'danger')
            return redirect(url_for('master.alumnos_admin'))
        try:
            grupo_int = int(grupo)
        except (TypeError, ValueError):
            flash('El grupo debe ser un número entero.', 'danger')
            return redirect(url_for('master.alumnos_admin'))

        alumno = Alumno(nombre=nombre, cmp=cmp, grupo=grupo_int)
        if id_val:
            if not id_val.isdigit():
                flash('El código del alumno debe ser numérico.', 'danger')
                return redirect(url_for('master.alumnos_admin'))
            alumno.id = int(id_val)
        db.session.add(alumno)
        try:
            db.session.flush()
            enroll_student(examen_activo, alumno)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('No se pudo agregar el alumno: el código o identificador ya existe.', 'danger')
            return redirect(url_for('master.alumnos_admin'))
        flash('Alumno agregado exitosamente.', 'success')
        return redirect(url_for('master.alumnos_admin'))
        
    alumnos = (
        Alumno.query
        .join(ExamenAlumno)
        .filter(ExamenAlumno.examen_id == examen_activo.id)
        .order_by(Alumno.grupo, Alumno.nombre)
        .all()
    )
    return render_template('master_alumnos.html', alumnos=alumnos, examen_activo=examen_activo)

@master_bp.route('/alumnos/editar/<int:id>', methods=['GET', 'POST'])
def alumnos_editar(id):
    alumno = Alumno.query.get_or_404(id)
    examen_activo = get_active_exam()
    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        grupo = (request.form.get('grupo') or '').strip()
        cmp = (request.form.get('cmp') or '').strip() or None
        if not nombre:
            flash('El nombre del alumno es obligatorio.', 'danger')
            return redirect(url_for('master.alumnos_editar', id=id))
        try:
            grupo_int = int(grupo)
        except (TypeError, ValueError):
            flash('El grupo debe ser un número entero.', 'danger')
            return redirect(url_for('master.alumnos_editar', id=id))

        alumno.nombre = nombre
        alumno.grupo = grupo_int
        alumno.cmp = cmp

        db.session.commit()
        flash('Alumno actualizado correctamente.', 'success')
        return redirect(url_for('master.alumnos_admin'))
        
    alumnos = (
        Alumno.query
        .join(ExamenAlumno)
        .filter(ExamenAlumno.examen_id == examen_activo.id)
        .order_by(Alumno.grupo, Alumno.nombre)
        .all()
    )
    return render_template('master_alumnos.html', alumnos=alumnos, alumno_edit=alumno, examen_activo=examen_activo)

@master_bp.route('/alumnos/eliminar/<int:id>', methods=['POST'])
def alumnos_eliminar(id):
    alumno = Alumno.query.get_or_404(id)
    examen_activo = get_active_exam(create_if_missing=False)
    if examen_activo:
        inscripcion = ExamenAlumno.query.filter_by(examen_id=examen_activo.id, alumno_id=alumno.id).first()
        if inscripcion:
            db.session.delete(inscripcion)
    db.session.commit()
    flash('Alumno retirado del examen activo. Su histórico no fue borrado.', 'success')
    return redirect(url_for('master.alumnos_admin'))

@master_bp.route('/alumnos/plantilla')
def alumnos_plantilla():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['codigo', 'nombres', 'apellidos', 'grupo', 'cmp', 'correo', 'documento', 'observaciones'])
    writer.writerow(['1001', 'Ana María', 'Pérez Ramos', '1', 'CMP12345', 'ana@example.com', 'DNI12345678', 'Ejemplo opcional'])
    data = '\ufeff' + output.getvalue()
    return Response(
        data,
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=plantilla_alumnos_teleecoe.csv'}
    )

@master_bp.route('/alumnos/plantilla-xlsx')
def alumnos_plantilla_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = 'alumnos'
    ws.append(['codigo', 'nombres', 'apellidos', 'grupo', 'cmp', 'correo', 'documento', 'observaciones'])
    ws.append(['1001', 'Ana María', 'Pérez Ramos', 1, 'CMP12345', 'ana@example.com', 'DNI12345678', 'Ejemplo opcional'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=plantilla_alumnos_teleecoe.xlsx'}
    )

def read_student_rows(archivo):
    filename = (archivo.filename or '').lower()
    raw = archivo.read()
    if filename.endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or '').strip().lower() for h in rows[0]]
        result = []
        for row in rows[1:]:
            result.append({headers[i]: '' if value is None else str(value).strip() for i, value in enumerate(row) if i < len(headers)})
        return result

    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return []
    return [{str(k).strip().lower(): (v or '').strip() for k, v in row.items() if k is not None} for row in reader]

@master_bp.route('/alumnos/importar', methods=['POST'])
def alumnos_importar():
    examen_activo = get_active_exam()
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Selecciona un archivo Excel (.xlsx) o CSV.', 'danger')
        return redirect(url_for('master.alumnos_admin'))

    rows = read_student_rows(archivo)
    if not rows:
        flash('El archivo no tiene encabezados.', 'danger')
        return redirect(url_for('master.alumnos_admin'))

    creados = actualizados = inscritos = errores = 0
    for idx, norm in enumerate(rows, start=2):
        codigo = norm.get('codigo') or norm.get('id') or norm.get('dni') or norm.get('documento')
        nombres = norm.get('nombres') or norm.get('nombre') or ''
        apellidos = norm.get('apellidos') or norm.get('apellido') or ''
        nombre_completo = (norm.get('nombre completo') or f'{nombres} {apellidos}'.strip()).strip()
        grupo_txt = norm.get('grupo') or '1'
        cmp = norm.get('cmp') or norm.get('codigo_cmp') or None
        observaciones = norm.get('observaciones') or None

        if not nombre_completo:
            errores += 1
            continue
        try:
            grupo = int(float(grupo_txt)) if grupo_txt else 1
        except ValueError:
            grupo = 1

        alumno = None
        if codigo and str(codigo).isdigit():
            alumno = Alumno.query.get(int(codigo))
        if not alumno and cmp:
            alumno = Alumno.query.filter_by(cmp=cmp).first()

        if alumno:
            alumno.nombre = nombre_completo
            alumno.grupo = grupo
            if cmp:
                alumno.cmp = cmp
            actualizados += 1
        else:
            alumno = Alumno(nombre=nombre_completo, cmp=cmp, grupo=grupo)
            if codigo and str(codigo).isdigit():
                alumno.id = int(codigo)
            db.session.add(alumno)
            try:
                db.session.flush()
            except Exception:
                db.session.rollback()
                alumno = Alumno(nombre=nombre_completo, cmp=cmp, grupo=grupo)
                db.session.add(alumno)
                db.session.flush()
            creados += 1

        if not ExamenAlumno.query.filter_by(examen_id=examen_activo.id, alumno_id=alumno.id).first():
            enroll_student(examen_activo, alumno, observaciones=observaciones)
            inscritos += 1

    db.session.commit()
    flash(f'Importación completada: {creados} creados, {actualizados} actualizados, {inscritos} inscritos, {errores} filas omitidas.', 'success' if errores == 0 else 'warning')
    return redirect(url_for('master.alumnos_admin'))

@master_bp.route('/estaciones', methods=['GET', 'POST'])
def estaciones_admin():
    if request.method == 'POST':
        id_val = (request.form.get('id') or '').strip()
        nombre = (request.form.get('nombre') or '').strip()
        orden = (request.form.get('orden') or '').strip()
        activa = request.form.get('activa') == '1'

        if not id_val or not nombre:
            flash('El ID y nombre de la estación son obligatorios.', 'danger')
            return redirect(url_for('master.estaciones_admin'))
        try:
            orden_int = int(orden)
        except (TypeError, ValueError):
            flash('El orden de la estación debe ser un número entero.', 'danger')
            return redirect(url_for('master.estaciones_admin'))

        estacion = Estacion(id=id_val, nombre=nombre, orden=orden_int, activa=activa)
        db.session.add(estacion)
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash('No se pudo agregar la estación: el ID ya existe.', 'danger')
            return redirect(url_for('master.estaciones_admin'))
        flash('Estación agregada.', 'success')
        return redirect(url_for('master.estaciones_admin'))
        
    estaciones = Estacion.query.order_by(Estacion.orden).all()
    return render_template('master_estaciones.html', estaciones=estaciones)

@master_bp.route('/estaciones/editar/<id>', methods=['GET', 'POST'])
def estaciones_editar(id):
    estacion = Estacion.query.get_or_404(id)
    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        orden = (request.form.get('orden') or '').strip()
        activa = request.form.get('activa') == '1'
        if not nombre:
            flash('El nombre de la estación es obligatorio.', 'danger')
            return redirect(url_for('master.estaciones_editar', id=id))
        try:
            orden_int = int(orden)
        except (TypeError, ValueError):
            flash('El orden de la estación debe ser un número entero.', 'danger')
            return redirect(url_for('master.estaciones_editar', id=id))

        estacion.nombre = nombre
        estacion.orden = orden_int
        estacion.activa = activa
        db.session.commit()
        flash('Estación actualizada.', 'success')
        return redirect(url_for('master.estaciones_admin'))
    estaciones = Estacion.query.order_by(Estacion.orden).all()
    return render_template('master_estaciones.html', estaciones=estaciones, estacion_edit=estacion)

@master_bp.route('/estaciones/eliminar/<id>', methods=['POST'])
def estaciones_eliminar(id):
    estacion = Estacion.query.get_or_404(id)

    try:
        evaluaciones = Evaluacion.query.filter_by(estacion_id=estacion.id).all()
        evaluacion_ids = [evaluacion.id for evaluacion in evaluaciones]
        if evaluacion_ids:
            EvaluacionDetalle.query.filter(EvaluacionDetalle.evaluacion_id.in_(evaluacion_ids)).delete(synchronize_session=False)
            Evaluacion.query.filter(Evaluacion.id.in_(evaluacion_ids)).delete(synchronize_session=False)

        for categoria in list(estacion.categorias):
            for criterio in list(categoria.criterios):
                db.session.delete(criterio)
            db.session.delete(categoria)

        db.session.delete(estacion)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('No se pudo eliminar la estación porque tiene datos relacionados.', 'danger')
        return redirect(url_for('master.estaciones_admin'))

    flash('Estación eliminada junto con su rúbrica y evaluaciones asociadas.', 'success')
    return redirect(url_for('master.estaciones_admin'))

@master_bp.route('/estaciones/toggle/<id>', methods=['POST'])
def estaciones_toggle(id):
    estacion = Estacion.query.get_or_404(id)
    estacion.activa = not estacion.esta_activa
    db.session.commit()
    estado = 'activada' if estacion.esta_activa else 'desactivada'
    flash(f'Estación {estado}: {estacion.nombre}.', 'success')
    return redirect(url_for('master.estaciones_admin'))

@master_bp.route('/estaciones/<id>/constructor', methods=['GET'])
def estaciones_constructor(id):
    estacion = Estacion.query.get_or_404(id)
    return render_template('master_constructor.html', estacion=estacion)

@master_bp.route('/categorias/nueva/<id_estacion>', methods=['POST'])
def categorias_nueva(id_estacion):
    estacion = Estacion.query.get_or_404(id_estacion)
    nombre = request.form.get('nombre')
    orden = request.form.get('orden', type=int, default=1)
    tipo = request.form.get('tipo', default='normal')
    
    categoria = Categoria(estacion_id=id_estacion, nombre=nombre, orden=orden, tipo=tipo)
    db.session.add(categoria)
    db.session.commit()
    flash('Categoría añadida exitosamente.', 'success')
    return redirect(url_for('master.estaciones_constructor', id=id_estacion))

@master_bp.route('/categorias/eliminar/<id>', methods=['POST'])
def categorias_eliminar(id):
    categoria = Categoria.query.get_or_404(id)
    id_estacion = categoria.estacion_id
    
    for criterio in categoria.criterios:
        db.session.delete(criterio)
        
    db.session.delete(categoria)
    db.session.commit()
    flash('Categoría eliminada.', 'success')
    return redirect(url_for('master.estaciones_constructor', id=id_estacion))

@master_bp.route('/criterios/nuevo/<id_categoria>', methods=['POST'])
def criterios_nuevo(id_categoria):
    categoria = Categoria.query.get_or_404(id_categoria)
    texto = (request.form.get('texto') or '').strip()
    requested_id = request.form.get('id') or texto
    id_val = unique_criterio_id(requested_id)
    puntos = request.form.get('puntos', type=float, default=1.0)
    tipo = request.form.get('tipo', default='checkbox')
    opciones = request.form.get('opciones', default='') if tipo == 'rango' else None

    if not texto:
        flash('El texto del criterio es obligatorio.', 'danger')
        return redirect(url_for('master.estaciones_constructor', id=categoria.estacion_id))
    
    criterio = Criterio(id=id_val, categoria_id=id_categoria, texto=texto, puntos=puntos, tipo=tipo, opciones=opciones)
    db.session.add(criterio)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        id_val = unique_criterio_id(f'{categoria.estacion_id}_{texto}')
        criterio.id = id_val
        db.session.add(criterio)
        db.session.commit()

    if slugify_identifier(requested_id) != id_val:
        flash(f'Criterio añadido exitosamente con ID único: {id_val}.', 'success')
    else:
        flash('Criterio añadido exitosamente.', 'success')
    return redirect(url_for('master.estaciones_constructor', id=categoria.estacion_id))

@master_bp.route('/criterios/eliminar/<id>', methods=['POST'])
def criterios_eliminar(id):
    criterio = Criterio.query.get_or_404(id)
    id_estacion = criterio.categoria.estacion_id
    db.session.delete(criterio)
    db.session.commit()
    flash('Criterio eliminado.', 'success')
    return redirect(url_for('master.estaciones_constructor', id=id_estacion))

@master_bp.route('/exportar')
def exportar_csv():
    examen_activo = get_active_exam()
    alumnos = (
        Alumno.query
        .join(ExamenAlumno)
        .filter(ExamenAlumno.examen_id == examen_activo.id)
        .order_by(Alumno.grupo, Alumno.nombre)
        .all()
    )
    estaciones = Estacion.query.filter_by(activa=True).order_by(Estacion.orden).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Panel de control'

    encabezados = ['ID', 'Alumno', 'Grupo']
    for est in estaciones:
        encabezados.extend([f'E{est.orden} {est.nombre} - Evaluado', f'E{est.orden} {est.nombre} - Puntaje'])
    ws.append(encabezados)

    for alumno in alumnos:
        evaluaciones = {
            e.estacion_id: e
            for e in alumno.evaluaciones
            if e.examen_id == examen_activo.id
        }
        fila = [alumno.id, alumno.nombre, alumno.grupo]
        for est in estaciones:
            evaluacion = evaluaciones.get(est.id)
            fila.extend(['Si' if evaluacion else 'No', round(evaluacion.puntaje_total, 1) if evaluacion else 0.0])
        ws.append(fila)

    style_export_sheet(ws)

    resumen = wb.create_sheet('Resumen')
    resumen.append(['Campo', 'Valor'])
    resumen.append(['Examen activo', examen_activo.nombre])
    resumen.append(['Alumnos', len(alumnos)])
    resumen.append(['Estaciones activas', len(estaciones)])
    resumen.append(['Filas exportadas', len(alumnos)])
    style_export_sheet(resumen)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype=XLSX_MIMETYPE,
        headers={'Content-Disposition': 'attachment; filename=panel_control_calificaciones.xlsx'}
    )
