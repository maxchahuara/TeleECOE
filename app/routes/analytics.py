from flask import Blueprint, Response, jsonify, render_template, request
from app.models import Examen, ExamenAlumno, Alumno, Estacion, Categoria, Criterio, Evaluacion, EvaluacionDetalle
from extensions import db
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

analytics_bp = Blueprint('analytics', __name__)
XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def current_exam_filter():
    examen_id = request.args.get('examen_id', type=int)
    if examen_id:
        return examen_id
    activo = Examen.query.filter_by(estado='activo').order_by(Examen.id.desc()).first()
    return activo.id if activo else None

def get_exam_groups(examen_id):
    query = db.session.query(Alumno.grupo)
    if examen_id:
        query = query.join(ExamenAlumno).filter(ExamenAlumno.examen_id == examen_id)
    return [
        grupo
        for grupo, in query.filter(Alumno.grupo.isnot(None)).distinct().order_by(Alumno.grupo).all()
    ]

def filtered_evaluations_query(examen_id, estacion_id=None, grupo=None):
    query = Evaluacion.query
    if examen_id:
        query = query.filter_by(examen_id=examen_id)
    if estacion_id:
        query = query.filter_by(estacion_id=estacion_id)
    if grupo is not None:
        query = query.join(Alumno).filter(Alumno.grupo == grupo)
    return query

def analytics_summary(examen_id, grupo_filter=None):
    alumnos_query = Alumno.query
    if examen_id:
        alumnos_query = alumnos_query.join(ExamenAlumno).filter(ExamenAlumno.examen_id == examen_id)
    if grupo_filter is not None:
        alumnos_query = alumnos_query.filter(Alumno.grupo == grupo_filter)

    promedio_query = db.session.query(func.avg(Evaluacion.puntaje_total))
    if examen_id:
        promedio_query = promedio_query.filter(Evaluacion.examen_id == examen_id)
    if grupo_filter is not None:
        promedio_query = promedio_query.join(Alumno).filter(Alumno.grupo == grupo_filter)

    return {
        'total_alumnos': alumnos_query.distinct().count(),
        'total_estaciones': Estacion.query.filter_by(activa=True).count(),
        'total_evaluaciones': filtered_evaluations_query(examen_id, grupo=grupo_filter).count(),
        'promedio_global': round(promedio_query.scalar() or 0.0, 2),
    }

def station_metrics(examen_id, grupo_filter=None):
    resultados = []
    estaciones = Estacion.query.filter_by(activa=True).order_by(Estacion.orden).all()
    for est in estaciones:
        evals = filtered_evaluations_query(examen_id, est.id, grupo_filter).all()
        n_evals = len(evals)
        if n_evals == 0:
            resultados.append({
                'estacion_id': est.id,
                'nombre': est.nombre,
                'orden': est.orden,
                'n_evaluaciones': 0,
                'promedio': 0,
                'maximo': 0,
                'minimo': 0,
            })
            continue

        puntajes = [e.puntaje_total for e in evals]
        resultados.append({
            'estacion_id': est.id,
            'nombre': est.nombre,
            'orden': est.orden,
            'n_evaluaciones': n_evals,
            'promedio': round(sum(puntajes) / n_evals, 2),
            'maximo': round(max(puntajes), 2),
            'minimo': round(min(puntajes), 2),
        })
    return resultados

def item_metrics(examen_id, estacion_filter=None, grupo_filter=None):
    evaluaciones = filtered_evaluations_query(examen_id, estacion_filter, grupo_filter).all()
    eval_ids = [e.id for e in evaluaciones]
    if not eval_ids:
        return []

    criterios_query = Criterio.query.join(Categoria).join(Estacion)
    if estacion_filter:
        criterios_query = criterios_query.filter(Estacion.id == estacion_filter)

    criterios = criterios_query.order_by(Estacion.orden, Categoria.orden, Criterio.id).all()
    detalles = EvaluacionDetalle.query.filter(EvaluacionDetalle.evaluacion_id.in_(eval_ids)).all()
    detalles_por_criterio = {}
    for d in detalles:
        detalles_por_criterio.setdefault(d.criterio_id, []).append(d)

    resultados = []
    for c in criterios:
        evals_esta_estacion = [e for e in evaluaciones if e.estacion_id == c.categoria.estacion_id]
        n_intentos = len(evals_esta_estacion)
        if n_intentos == 0:
            continue

        aciertos = 0
        fallos = 0
        d_list = detalles_por_criterio.get(c.id, [])
        if c.tipo != 'rango':
            aciertos = len(d_list)
            fallos = n_intentos - aciertos
        else:
            for d in d_list:
                try:
                    valor_num = float(d.valor_registrado)
                    if valor_num > 0.0:
                        aciertos += 1
                    else:
                        fallos += 1
                except Exception:
                    fallos += 1
            fallos += n_intentos - len(d_list)

        acierto_rate = (aciertos / n_intentos) * 100 if n_intentos > 0 else 0
        fallo_rate = (fallos / n_intentos) * 100 if n_intentos > 0 else 0
        resultados.append({
            'criterio_id': c.id,
            'texto': c.texto,
            'categoria': c.categoria.nombre,
            'estacion': c.categoria.estacion.nombre,
            'estacion_id': c.categoria.estacion.id,
            'n_intentos': n_intentos,
            'aciertos': aciertos,
            'fallos': fallos,
            'acierto_rate': round(acierto_rate, 1),
            'fallo_rate': round(fallo_rate, 1),
        })

    return sorted(resultados, key=lambda x: x['fallo_rate'], reverse=True)

def style_sheet(ws):
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
    for column_cells in ws.columns:
        width = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 60)
    ws.freeze_panes = 'A2'

@analytics_bp.route('/')
def dashboard_view():
    """
    Renderiza la vista principal del dashboard analítico.
    """
    examen_id = current_exam_filter()
    estaciones = Estacion.query.filter_by(activa=True).order_by(Estacion.orden).all()
    grupos = get_exam_groups(examen_id)
    return render_template('analytics_dashboard.html', estaciones=estaciones, grupos=grupos)


@analytics_bp.route('/api/summary')
def api_summary():
    """
    KPIs globales del sistema.
    """
    grupo_filter = request.args.get('grupo', type=int)
    return jsonify(analytics_summary(current_exam_filter(), grupo_filter))

@analytics_bp.route('/api/items')
def api_items():
    """
    Devuelve la matriz de ítems con sus métricas calculadas (Tasa de fallo, acierto, etc).
    Soporta filtrado por ?estacion_id= y ?grupo=
    """
    estacion_filter = request.args.get('estacion_id')
    grupo_filter = request.args.get('grupo', type=int)

    return jsonify(item_metrics(current_exam_filter(), estacion_filter, grupo_filter))


@analytics_bp.route('/api/stations')
def api_stations():
    """
    Retorna la distribución de puntajes e información global agrupada por estación.
    """
    grupo_filter = request.args.get('grupo', type=int)
    return jsonify(station_metrics(current_exam_filter(), grupo_filter))

@analytics_bp.route('/export')
def export_csv():
    """
    Descarga el dashboard analítico como Excel: KPIs, tablas y gráficos.
    """
    estacion_filter = request.args.get('estacion_id')
    grupo_filter = request.args.get('grupo', type=int)
    examen_id = current_exam_filter()

    wb = Workbook()

    ws = wb.active
    ws.title = 'KPIs'
    ws.append(['Indicador', 'Valor'])
    resumen = analytics_summary(examen_id, grupo_filter)
    ws.append(['Evaluaciones totales', resumen['total_evaluaciones']])
    ws.append(['Promedio global', resumen['promedio_global']])
    ws.append(['Estaciones activas', resumen['total_estaciones']])
    ws.append(['Alumnos evaluados', resumen['total_alumnos']])
    style_sheet(ws)

    ws_est = wb.create_sheet('Tabla estaciones')
    ws_est.append(['Orden', 'Estacion', 'Evaluaciones', 'Promedio', 'Maximo', 'Minimo'])
    for row in station_metrics(examen_id, grupo_filter):
        ws_est.append([
            row['orden'],
            row['nombre'],
            row['n_evaluaciones'],
            row['promedio'],
            row['maximo'],
            row['minimo'],
        ])
    style_sheet(ws_est)

    ws_items = wb.create_sheet('Tabla items')
    ws_items.append([
        'Estacion', 'Categoria', 'Criterio ID', 'Item evaluado',
        'Intentos', 'Aciertos', 'Fallos', 'Tasa acierto %', 'Tasa fallo %', 'Estado'
    ])
    items = item_metrics(examen_id, estacion_filter, grupo_filter)
    for row in items:
        ws_items.append([
            row['estacion'],
            row['categoria'],
            row['criterio_id'],
            row['texto'],
            row['n_intentos'],
            row['aciertos'],
            row['fallos'],
            row['acierto_rate'],
            row['fallo_rate'],
            'Critico' if row['fallo_rate'] > 50 else 'Optimo',
        ])
    style_sheet(ws_items)

    ws_charts = wb.create_sheet('Graficos')
    ws_charts['A1'] = 'Promedio de puntaje por estacion'
    ws_charts['A20'] = 'Top 5 items criticos por tasa de fallo'

    if ws_est.max_row > 1:
        chart = BarChart()
        chart.title = 'Promedio de puntaje por estacion'
        chart.y_axis.title = 'Promedio'
        chart.x_axis.title = 'Estacion'
        data = Reference(ws_est, min_col=4, min_row=1, max_row=ws_est.max_row)
        cats = Reference(ws_est, min_col=2, min_row=2, max_row=ws_est.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_charts.add_chart(chart, 'A2')

    if ws_items.max_row > 1:
        chart = BarChart()
        chart.type = 'bar'
        chart.title = 'Top 5 items criticos por tasa de fallo'
        chart.y_axis.title = 'Item'
        chart.x_axis.title = 'Tasa fallo %'
        top_count = min(5, len(items))
        data = Reference(ws_items, min_col=9, min_row=1, max_row=top_count + 1)
        cats = Reference(ws_items, min_col=4, min_row=2, max_row=top_count + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_charts.add_chart(chart, 'A21')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype=XLSX_MIMETYPE,
        headers={"Content-Disposition": "attachment; filename=dashboard_analitico_teleecoe.xlsx"}
    )
